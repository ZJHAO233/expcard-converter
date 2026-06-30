"""
DCS Converter - Excel试验条件表转Markdown工具
"""

import re
import openpyxl


class DCSConverter:

    SKIP_HEADERS = ['序号', '条件确认']
    SECTION_HEADERS = ['试验条件', '试验恢复', '结论', '存在问题']

    # 逻辑运算符字典: 键为Excel中的值, 值为输出时使用的符号
    # 可在此处补充新的逻辑运算符
    LOGIC_OPERATORS = {
        '与': '且',
        '且': '且',
        '或': '或',
        '或取反': '或取反',
    }

    # 特殊分隔符字典: 键为特殊分隔符的值, 值为处理方式描述
    # 用于模糊检测特殊分隔符, 在默认逻辑分隔符判断之前进行检测
    SPECIAL_SEPARATORS = {
        # 示例: '特殊值': '处理方式说明',
    }

    # 所有可能的逻辑分隔符（用于识别）
    LOGIC_SEPARATORS = list(LOGIC_OPERATORS.keys())

    def __init__(self):
        self.output = []
        self.in_content_area = False
        self.skip_section = False
        self.current_sub_title_logic = None
        self.processed_rows = set()
        self.special_separator_positions = set()  # 记录特殊分隔符匹配的行列号

    def convert(self, file_path, sheet_index=0):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[sheet_index]

        rows = []
        for r in range(1, ws.max_row + 1):
            row = []
            for c in range(1, min(ws.max_column + 1, 8)):
                val = ws.cell(r, c).value
                row.append(str(val).strip() if val else '')
            rows.append(row)

        self.output = []
        self.in_content_area = False
        self.skip_section = False
        self.current_sub_title_logic = None
        self.processed_rows = set()
        self.special_separator_positions = set()

        for i, row in enumerate(rows):
            if i in self.processed_rows:
                continue
            self._process_row(row, rows, i)

        return '\n'.join(self.output)

    def _is_special_separator(self, text, row_index=None, col_index=None):
        """检测是否为特殊分隔符, 如果匹配则记录行列号并返回True"""
        if text in self.SPECIAL_SEPARATORS:
            if row_index is not None and col_index is not None:
                self.special_separator_positions.add((row_index, col_index))
            return True
        return False

    def _process_row(self, row, all_rows, index):
        a = row[0] if len(row) > 0 else ''
        b = row[1] if len(row) > 1 else ''
        c = row[2] if len(row) > 2 else ''

        if not a and not b:
            return
        if self._is_header_row(row):
            return

        # 二级标题
        if self._is_chinese_number_title(a):
            self.output.append('')
            self.output.append('## ' + a)
            self.in_content_area = False
            self.skip_section = False
            return

        # 段落标题
        if a in self.SECTION_HEADERS:
            if a == '试验条件':
                self.skip_section = True
            self.in_content_area = False
            return

        # 进入内容区域
        if a == '试验内容':
            self.in_content_area = True
            self.skip_section = False
            return

        if self.skip_section:
            return

        # 三级标题
        if self._is_sub_title(a):
            # 统一括号格式为全角
            normalized = a.replace('(', '（').replace(')', '）')
            self.output.append('')
            self.output.append('### ' + normalized)
            self.output.append('')
            self.current_sub_title_logic = self._extract_title_logic(normalized)
            return

        # 子列一级
        if a.isdigit() and '.' not in a:
            self._process_level1(a, b, c, row, all_rows, index)
            return

        # 子列二级
        if self._is_sub_item(a):
            self._process_level2(a, b, c, row, all_rows, index)
            return

    def _is_header_row(self, row):
        a, b = row[0], row[1]
        if a in self.SKIP_HEADERS:
            return True
        if a == '序号' and ('条件' in b or '内容' in b):
            return True
        return False

    def _is_chinese_number_title(self, text):
        return bool(re.match(r'^[一二三四五六七八九十]+、', text))

    def _is_sub_title(self, text):
        return bool(re.match(r'^\d+\.[\u4e00-\u9fa5]', text))

    def _is_sub_item(self, text):
        return bool(re.match(r'^\d+\.\d+$', text))

    def _extract_title_logic(self, title):
        match = re.search(r'[（(]([与或])[）)]', title)
        return match.group(1) if match else None

    def _is_logic_separator(self, text):
        return text in self.LOGIC_SEPARATORS

    def _get_logic_output(self, logic):
        """获取逻辑运算符的输出格式"""
        return self.LOGIC_OPERATORS.get(logic, logic)

    def _get_content_from_row(self, row, start_col):
        for i in range(start_col, len(row)):
            val = row[i]
            if val and not self._is_logic_separator(val):
                return val
        return None

    def _process_level1(self, a, b, c, row, all_rows, index, use_seq_num=False):
        # 先检测特殊分隔符（在默认逻辑分隔符判断之前）
        if b:
            self._is_special_separator(b, row_index=index, col_index=1)

        # 判断是否有子项
        has_children = False
        child_logic = None
        if index + 1 < len(all_rows):
            next_a = all_rows[index + 1][0]
            if self._is_sub_item(next_a) and next_a.startswith(a + '.'):
                has_children = True
                # 从第一个子项的B列获取逻辑
                child_logic = all_rows[index + 1][1] if self._is_logic_separator(all_rows[index + 1][1]) else None

        # 获取内容
        content = self._get_content_from_row(row, start_col=1)
        if not content:
            content = b if b and not self._is_logic_separator(b) else ''

        # 判断是否输出逻辑
        logic_str = ''
        if has_children and child_logic:
            # 只有当子列一级自己的B列逻辑与三级标题逻辑一致时才忽略
            # 如果子列一级B列不是逻辑分隔符，则使用子项的逻辑
            if self._is_logic_separator(b):
                # 子列一级B列是逻辑分隔符，检查是否与三级标题一致
                if b != self.current_sub_title_logic:
                    logic_str = '（' + b + '）'
            else:
                # 子列一级B列不是逻辑分隔符，使用子项的逻辑
                logic_str = '（' + child_logic + '）'
        
        # 处理或取反
        if b and '或取反' in b:
            logic_str = '（或取反）'

        # 使用顺序编号还是原始A列编号
        if use_seq_num:
            self.output.append(str(self.item_counter) + '. ' + content + logic_str)
        else:
            self.output.append(a + '. ' + content + logic_str)

    def _process_level2(self, a, b, c, row, all_rows, index):
        # 子列二级的列含义:
        # B列: 可能是逻辑分隔符（传递给父级）或内容
        # C列: 二级分组逻辑 或 三级条件逻辑（当D列也是逻辑分隔符时）或内容
        # D列: 二级分组内容 或 三级条件逻辑 或内容
        # E列: 三级条件内容

        # 获取各列的值
        b_val = b
        c_val = c
        d_val = row[3] if len(row) > 3 else ''
        e_val = row[4] if len(row) > 4 else ''

        # 先检测特殊分隔符（在默认逻辑分隔符判断之前）
        if b_val:
            self._is_special_separator(b_val, row_index=index, col_index=1)
        if c_val:
            self._is_special_separator(c_val, row_index=index, col_index=2)
        if d_val:
            self._is_special_separator(d_val, row_index=index, col_index=3)
        if e_val:
            self._is_special_separator(e_val, row_index=index, col_index=4)

        b_is_logic = self._is_logic_separator(b_val)
        c_is_logic = self._is_logic_separator(c_val)
        d_is_logic = self._is_logic_separator(d_val)

        # 情况1: C列和D列都是逻辑分隔符 → 三级条件处理
        # （不管B列是什么）
        if c_is_logic and d_is_logic:
            self._process_level3_group(a, c_val, d_val, e_val, row, all_rows, index)
            return

        # 情况2: C列是逻辑分隔符，D列不是逻辑分隔符 → 二级分组
        # （不管B列是什么）
        if c_is_logic and not d_is_logic:
            self._process_level2_group(a, c_val, d_val, row, all_rows, index)
            return

        # 情况3: B列是逻辑分隔符
        if b_is_logic:
            # C列有内容（不是逻辑分隔符）→ 输出C列内容
            if c_val and not c_is_logic:
                self.output.append('   - ' + c_val)
                return
            # C列也是逻辑分隔符 → 已在情况1/2处理
            # C列为空 → 跳过
            return

        # 情况4: 简单输出
        # 从B列开始找内容（跳过逻辑分隔符）
        content = None
        for col_idx in range(1, len(row)):
            val = row[col_idx]
            if val and not self._is_logic_separator(val):
                content = val
                break

        if content:
            self.output.append('   - ' + content)

    def _process_level3_group(self, a, c_logic, d_logic, first_content, row, all_rows, index):
        """处理三级条件分组: C列和D列都是逻辑分隔符"""
        # 收集D列逻辑下的所有内容
        level3_groups = []
        current_group_items = [first_content] if first_content else []
        current_d_logic = d_logic

        j = index + 1
        while j < len(all_rows):
            next_row = all_rows[j]
            next_a = next_row[0]

            # 检查是否是同级子项（有编号的）
            if self._is_sub_item(next_a) and next_a.startswith(a.split('.')[0] + '.'):
                next_c = next_row[2] if len(next_row) > 2 else ''
                next_d = next_row[3] if len(next_row) > 3 else ''
                next_e = next_row[4] if len(next_row) > 4 else ''

                # 如果C列有逻辑，说明是新的二级分组，结束当前处理
                if self._is_logic_separator(next_c):
                    break

                # 如果D列有逻辑，说明是新的三级分组
                if self._is_logic_separator(next_d):
                    # 保存当前分组
                    if current_group_items:
                        level3_groups.append((current_d_logic, current_group_items[:]))
                    current_group_items = []
                    current_d_logic = next_d
                    # 从E列获取内容
                    if next_e and not self._is_logic_separator(next_e):
                        current_group_items.append(next_e)
                        self.processed_rows.add(j)
                # D列有内容但不是逻辑分隔符
                elif next_d and not self._is_logic_separator(next_d):
                    # 保存当前分组
                    if current_group_items:
                        level3_groups.append((current_d_logic, current_group_items[:]))
                        current_group_items = []
                    # 将D列内容作为单独项目
                    level3_groups.append((None, [next_d]))
                    self.processed_rows.add(j)
                # D列为空，从E列获取内容加入当前分组
                elif next_e and not self._is_logic_separator(next_e):
                    current_group_items.append(next_e)
                    self.processed_rows.add(j)
                j += 1

            # 检查是否是无编号的延续行（A列为空）
            elif not next_a:
                next_c = next_row[2] if len(next_row) > 2 else ''
                next_d = next_row[3] if len(next_row) > 3 else ''
                next_e = next_row[4] if len(next_row) > 4 else ''

                # 如果C列有逻辑，结束
                if self._is_logic_separator(next_c):
                    break

                # 如果D列有逻辑，新的三级分组
                if self._is_logic_separator(next_d):
                    if current_group_items:
                        level3_groups.append((current_d_logic, current_group_items[:]))
                    current_group_items = []
                    current_d_logic = next_d
                    if next_e and not self._is_logic_separator(next_e):
                        current_group_items.append(next_e)
                        self.processed_rows.add(j)
                # D列有内容但不是逻辑分隔符
                elif next_d and not self._is_logic_separator(next_d):
                    if current_group_items:
                        level3_groups.append((current_d_logic, current_group_items[:]))
                        current_group_items = []
                    level3_groups.append((None, [next_d]))
                    self.processed_rows.add(j)
                # 延续行，从E列获取内容
                elif next_e and not self._is_logic_separator(next_e):
                    current_group_items.append(next_e)
                    self.processed_rows.add(j)
                j += 1
            else:
                break

        # 保存最后一个分组
        if current_group_items:
            level3_groups.append((current_d_logic, current_group_items))

        # 构建三级条件结果
        if level3_groups:
            result_parts = []
            for logic, items in level3_groups:
                if logic is None:
                    # 单独项目，直接输出
                    result_parts.append(items[0])
                else:
                    output_logic = self._get_logic_output(logic)
                    separator = ' ' + output_logic + ' '
                    group_str = separator.join(items)
                    if len(items) > 1:
                        result_parts.append('（' + group_str + '）')
                    else:
                        result_parts.append(group_str)
            result = ' 或 '.join(result_parts)
            self.output.append('   - ' + result)

    def _process_level2_group(self, a, c_logic, first_content, row, all_rows, index):
        """处理二级分组: C列是逻辑分隔符，D列不是"""
        group_items = [first_content] if first_content else []
        continuation_logic = c_logic

        j = index + 1
        while j < len(all_rows):
            next_row = all_rows[j]
            next_a = next_row[0]

            # 检查是否是同级子项（有编号的）
            if self._is_sub_item(next_a) and next_a.startswith(a.split('.')[0] + '.'):
                next_c = next_row[2] if len(next_row) > 2 else ''
                next_d = next_row[3] if len(next_row) > 3 else ''

                # 如果C列有逻辑，说明是新的分组，结束当前处理
                if self._is_logic_separator(next_c):
                    break

                # 如果D列有逻辑，说明是新的三级条件，结束当前处理
                if self._is_logic_separator(next_d):
                    break

                # 从C列开始找内容
                next_content = None
                for col_idx in range(2, len(next_row)):
                    val = next_row[col_idx]
                    if val and not self._is_logic_separator(val):
                        next_content = val
                        break

                if next_content:
                    group_items.append(next_content)
                    self.processed_rows.add(j)
                j += 1

            # 检查是否是无编号的延续行（A列为空）
            elif not next_a:
                next_c = next_row[2] if len(next_row) > 2 else ''
                next_d = next_row[3] if len(next_row) > 3 else ''

                # 如果C列有逻辑，结束
                if self._is_logic_separator(next_c):
                    break

                # 如果D列有逻辑，结束
                if self._is_logic_separator(next_d):
                    break

                # 从C列开始找内容
                next_content = None
                for col_idx in range(2, len(next_row)):
                    val = next_row[col_idx]
                    if val and not self._is_logic_separator(val):
                        next_content = val
                        break

                if next_content:
                    group_items.append(next_content)
                    self.processed_rows.add(j)
                j += 1
            else:
                break

        # 输出分组结果
        if len(group_items) > 1:
            result = group_items[0]
            output_logic = self._get_logic_output(continuation_logic)
            separator = ' ' + output_logic + ' '
            for item in group_items[1:]:
                result = result + separator + item
            self.output.append('   - ' + result)
        elif group_items:
            self.output.append('   - ' + group_items[0])

    # ==================== New Mode Methods ====================

    def convert_new(self, file_path, sheet_index=0):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[sheet_index]

        rows = []
        for r in range(1, ws.max_row + 1):
            row = []
            for c in range(1, min(ws.max_column + 1, 8)):
                val = ws.cell(r, c).value
                row.append(str(val).strip() if val else '')
            rows.append(row)

        self.output = []
        self.in_content_area = False
        self.skip_section = False
        self.current_sub_title_logic = None
        self.processed_rows = set()
        self.special_separator_positions = set()
        self.current_sheet_name = ws.title
        self.item_counter = 0  # 顺序编号计数器

        for i, row in enumerate(rows):
            if i in self.processed_rows:
                continue
            self._process_new_row(row, rows, i)

        return '\n'.join(self.output)

    def _process_new_row(self, row, all_rows, index):
        a = row[0] if len(row) > 0 else ''
        b = row[1] if len(row) > 1 else ''

        if not a and not b:
            return
        if self._is_header_row(row):
            return

        # 二级标题
        if self._is_chinese_number_title(a):
            self.output.append('')
            self.output.append('## ' + a)
            self.in_content_area = False
            self.skip_section = False
            return

        # 段落标题
        if a in self.SECTION_HEADERS:
            if a == '试验条件':
                self.skip_section = True
            self.in_content_area = False
            return

        # 进入内容区域
        if a == '试验内容':
            self.in_content_area = True
            self.skip_section = False
            return

        if self.skip_section:
            return

        # 三级标题
        if self._is_sub_title(a):
            normalized = a.replace('(', '（').replace(')', '）')
            self.output.append('')
            self.output.append('### ' + normalized)
            self.output.append('')
            self.current_sub_title_logic = self._extract_title_logic(normalized)
            self.item_counter = 0  # 每个三级标题重新计数
            return

        # 纯数字序号 → 新模式处理
        if a.isdigit():
            self._process_new_item(a, row, all_rows, index)
            return

    def _process_new_item(self, a, row, all_rows, index):
        b = row[1] if len(row) > 1 else ''
        c = row[2] if len(row) > 2 else ''

        # 先检测特殊分隔符（在默认逻辑分隔符判断之前）
        if b:
            self._is_special_separator(b, row_index=index, col_index=1)
        if c:
            self._is_special_separator(c, row_index=index, col_index=2)

        # 检测下一行A列是否是x.y格式
        if index + 1 < len(all_rows):
            next_a = all_rows[index + 1][0]
            if self._is_sub_item(next_a):
                # 下一行是x.y格式 → old模式处理
                self.item_counter += 1
                self._process_level1(a, b, c, row, all_rows, index, use_seq_num=True)
                # 处理所有x.y子项
                j = index + 1
                while j < len(all_rows):
                    if j in self.processed_rows:
                        j += 1
                        continue
                    next_row = all_rows[j]
                    next_a_val = next_row[0]
                    if self._is_sub_item(next_a_val) and next_a_val.startswith(a + '.'):
                        self._process_level2(next_a_val, next_row[1] if len(next_row) > 1 else '',
                                            next_row[2] if len(next_row) > 2 else '', next_row, all_rows, j)
                        self.processed_rows.add(j)
                        j += 1
                    else:
                        break
                return

        # 下一行不是x.y格式 → new模式执行检测
        b_is_logic = self._is_logic_separator(b)

        # B列不是逻辑 → 一级（直接输出）
        if not b_is_logic:
            content = b if b else ''
            self.item_counter += 1
            self.output.append(str(self.item_counter) + '. ' + content)
            return

        # B列是逻辑 → 进入二级处理
        self.item_counter += 1
        result = self._process_new_level2(self.item_counter, b, row, all_rows, index)
        self.output.append(str(self.item_counter) + '. ' + result)

    def _process_new_level2(self, item_num, b_logic, row, all_rows, index):
        c = row[2] if len(row) > 2 else ''

        # 先检测特殊分隔符（在默认逻辑分隔符判断之前）
        if c:
            self._is_special_separator(c, row_index=index, col_index=2)

        c_is_logic = self._is_logic_separator(c)

        # 确定二级遍历范围：从当前行开始，向下数B列空行直到下一个B列逻辑
        end_index = self._find_end_index(all_rows, index, col=1)

        items = []
        j = index
        while j <= end_index:
            if j >= len(all_rows):
                break
            next_row = all_rows[j]
            next_c = next_row[2] if len(next_row) > 2 else ''

            # 先检测特殊分隔符
            if next_c:
                self._is_special_separator(next_c, row_index=j, col_index=2)

            if j == index:
                # 当前行：C列是逻辑 → 三级处理
                if c_is_logic:
                    level3_result = self._process_new_level3(item_num, c, next_row, all_rows, j)
                    items.append(level3_result)
                # 当前行：C列是内容
                elif c:
                    items.append(c)
            else:
                # 后续行：C列是逻辑 → 三级处理
                if self._is_logic_separator(next_c):
                    level3_result = self._process_new_level3(item_num, next_c, next_row, all_rows, j)
                    items.append(level3_result)
                    self.processed_rows.add(j)
                # C列是内容
                elif next_c:
                    items.append(next_c)
                    self.processed_rows.add(j)

            j += 1

        # 用B列逻辑连接所有项目
        output_logic = self._get_logic_output(b_logic)
        separator = ' ' + output_logic + ' '
        return separator.join(items)

    def _process_new_level3(self, item_num, c_logic, row, all_rows, index):
        d = row[3] if len(row) > 3 else ''

        # 先检测特殊分隔符（在默认逻辑分隔符判断之前）
        if d:
            self._is_special_separator(d, row_index=index, col_index=3)

        d_is_logic = self._is_logic_separator(d)

        # 确定三级遍历范围：从当前行开始，向下数C列空行直到下一个C列逻辑
        end_index = self._find_end_index(all_rows, index, col=2)

        items = []
        j = index
        while j <= end_index:
            if j >= len(all_rows):
                break
            next_row = all_rows[j]
            next_d = next_row[3] if len(next_row) > 3 else ''

            # 先检测特殊分隔符
            if next_d:
                self._is_special_separator(next_d, row_index=j, col_index=3)

            if j == index:
                # 当前行：D列是逻辑 → 四级处理
                if d_is_logic:
                    level4_result = self._process_new_level4(item_num, d, next_row, all_rows, j)
                    items.append(level4_result)
                # 当前行：D列是内容
                elif d:
                    items.append(d)
            else:
                # 后续行：D列是逻辑 → 四级处理
                if self._is_logic_separator(next_d):
                    level4_result = self._process_new_level4(item_num, next_d, next_row, all_rows, j)
                    items.append(level4_result)
                    self.processed_rows.add(j)
                # D列是内容
                elif next_d:
                    items.append(next_d)
                    self.processed_rows.add(j)

            j += 1

        # 用C列逻辑连接，三层输出用（）包裹
        output_logic = self._get_logic_output(c_logic)
        separator = ' ' + output_logic + ' '
        result = separator.join(items)
        if len(items) > 1:
            return '（' + result + '）'
        return result

    def _process_new_level4(self, item_num, d_logic, row, all_rows, index):
        e = row[4] if len(row) > 4 else ''

        # 先检测特殊分隔符（在默认逻辑分隔符判断之前）
        if e:
            self._is_special_separator(e, row_index=index, col_index=4)

        # 确定四级遍历范围：从当前行开始，向下数D列空行直到下一个D列逻辑
        end_index = self._find_end_index(all_rows, index, col=3)

        items = []
        j = index
        while j <= end_index:
            if j >= len(all_rows):
                break
            next_row = all_rows[j]
            next_e = next_row[4] if len(next_row) > 4 else ''

            # 先检测特殊分隔符
            if next_e:
                self._is_special_separator(next_e, row_index=j, col_index=4)

            if j == index:
                # 当前行：E列内容
                if e and not self._is_logic_separator(e):
                    items.append(e)
            else:
                # 后续行：E列内容
                if next_e and not self._is_logic_separator(next_e):
                    items.append(next_e)
                    self.processed_rows.add(j)

            j += 1

        # 用D列逻辑连接，四层输出用（）包裹
        output_logic = self._get_logic_output(d_logic)
        separator = ' ' + output_logic + ' '
        result = separator.join(items)
        if len(items) > 1:
            return '（' + result + '）'
        return result

    def _find_end_index(self, all_rows, start_index, col):
        """找到当前层的结束行索引：从start_index开始，向下数col列空行直到遇到下一个逻辑分隔符"""
        j = start_index + 1
        while j < len(all_rows):
            next_row = all_rows[j]
            val = next_row[col] if len(next_row) > col else ''

            # 遇到逻辑分隔符 → 结束
            if self._is_logic_separator(val):
                return j - 1

            j += 1

        return len(all_rows) - 1


def convert_sheet(file_path, sheet_index=0, mode='old'):
    converter = DCSConverter()
    if mode == 'new':
        return converter.convert_new(file_path, sheet_index)
    return converter.convert(file_path, sheet_index)


def convert_all_sheets(file_path, mode='old'):
    """合并所有sheet页输出，每个sheet页为一级标题"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    all_output = []

    for i, ws in enumerate(wb.worksheets):
        sheet_name = ws.title
        result = convert_sheet(file_path, i, mode)
        if result.strip():
            all_output.append(f'# {sheet_name}\n')
            all_output.append(result)
            all_output.append('')

    return '\n'.join(all_output)


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 3:
        print("Usage: python convert.py <mode> <excel_file> [sheet_index|all]")
        print("  mode: old | new")
        print("  sheet_index: 0-based index, or 'all' for all sheets")
        sys.exit(1)

    mode = sys.argv[1]
    file_path = sys.argv[2]
    sheet_param = sys.argv[3] if len(sys.argv) > 3 else '0'

    if mode not in ('old', 'new'):
        print(f"Error: Unknown mode '{mode}'. Use 'old' or 'new'.")
        sys.exit(1)

    try:
        if sheet_param == 'all':
            result = convert_all_sheets(file_path, mode)
        else:
            sheet_index = int(sheet_param)
            result = convert_sheet(file_path, sheet_index, mode)

        with open('output.md', 'w', encoding='utf-8') as f:
            f.write(result)
        print("Done! Output written to output.md")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
